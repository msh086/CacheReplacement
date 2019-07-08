import argparse
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.layout import Layout, ManualLayout

import re


def get_data(filename):
    full_name = "/home/find/d/huawei/" + filename
    xls_file = "/home/find/d/huawei/%s.xls" % filename
    wb = Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1, value='Cache size')
    sheet.cell(row=1, column=2, value="Line size")
    sheet.cell(row=1, column=3, value="ways")
    sheet.cell(row=1, column=4, value="policy")
    sheet.cell(row=1, column=5, value="Memory-->Cache")
    sheet.cell(row=1, column=6, value="Cache-->Memory")
    sheet.cell(row=1, column=7, value="Hit rate %")
    sheet.cell(row=1, column=8, value="traffic")
    reg = re.compile(
        r"cache_size: (\d+).+line_size: (\d+).+mapping ways (\d+)[\s\S]+?hit/miss rate: (\d+\.\d+)[\s\S]+?Cache Policy=======\n(.+)\n[\s\S]+?Memory --> Cache:\t(\d+\.\d+)[\s\S]+?Cache --> Memory:\t(\d+\.\d+)")
    with open(full_name, 'r')as fin:
        content = fin.read()
    result = reg.findall(content)
    if not result:
        print("有错误，没有找到结果")
    else:
        for j, node in enumerate(result):
            sheet.cell(row=j + 2, column=1, value=int(node[0]))
            sheet.cell(row=j + 2, column=2, value=int(node[1]))
            sheet.cell(row=j + 2, column=3, value=int(node[2]))
            sheet.cell(row=j + 2, column=4, value=node[4])
            sheet.cell(row=j + 2, column=5, value=float(node[5]))
            sheet.cell(row=j + 2, column=6, value=float(node[6]) / 1024)
            sheet.cell(row=j + 2, column=7, value=float(node[3]))
            sheet.cell(row=j + 2, column=8, value=float(node[5]) + float(node[6]) / 1024)
        wb.save(xls_file)


get_data("cjzc.best.config")
